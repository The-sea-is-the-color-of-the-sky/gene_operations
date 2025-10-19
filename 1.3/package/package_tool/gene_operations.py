import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------------- 基础工具 ----------------
def validate_columns(df, *cols, set_progress_status=None):
    if set_progress_status:
        set_progress_status("验证列名...")
    for c in cols:
        if c not in df.columns:
            raise ValueError(f"缺少列: {c}")

def build_bidirectional_map(df, col1, col2, set_progress_status=None):
    if set_progress_status:
        set_progress_status("构建双向映射...")
    a_to_b, b_to_a = {}, {}
    for a, b in zip(df[col1], df[col2]):
        if pd.notnull(a) and pd.notnull(b):
            a_to_b.setdefault(a,set()).add(b)
            b_to_a.setdefault(b,set()).add(a)
    return a_to_b, b_to_a

def recursive_search(keys, a_to_b, b_to_a, depth=3, fuzzy=False, set_progress_status=None, sub_progress_callback=None):
    found = set(keys)
    current = set(keys)
    for d in range(depth):
        next_found = set()
        total = len(current)
        for idx, key in enumerate(current):
            if fuzzy:
                next_found.update(
                    b for a in a_to_b if key in str(a) and a != key for b in a_to_b[a]
                )
                next_found.update(
                    a for b in b_to_a if key in str(b) and b != key for a in b_to_a[b]
                )
            else:
                next_found |= a_to_b.get(key,set())
                next_found |= b_to_a.get(key,set())
            if sub_progress_callback and total>0:
                sub_progress_callback(int((idx+1)/total*100))
        new = next_found - found
        if not new:
            break
        found |= new
        current = new
    return found - set(keys)

# ---------------- 核心功能 ----------------
def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b,
                              progress_callback=None, sub_progress_callback=None, set_progress_status=None, vertical=False):
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)
    validate_columns(df_a, gene_column_a, set_progress_status=set_progress_status)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)
    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)

    all_matches=[]
    all_flags=[]
    total=len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_flags.append([])
            if progress_callback:
                progress_callback(int((i+1)/total*100))
            continue

        exact = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=False,
                                 set_progress_status=set_progress_status, sub_progress_callback=sub_progress_callback)
        fuzzy = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=True,
                                 set_progress_status=set_progress_status, sub_progress_callback=sub_progress_callback)
        exact_list = list(exact)
        fuzzy_list = list(fuzzy - exact)
        matches = exact_list + fuzzy_list
        flags = [False]*len(exact_list) + [True]*len(fuzzy_list)
        all_matches.append(matches)
        all_flags.append(flags)
        if progress_callback:
            progress_callback(int((i+1)/total*100))
        if set_progress_status and (i%10==0 or i==total-1):
            set_progress_status(f"处理进度：{i+1}/{total}")

    # 写入 Excel
    wb = Workbook()
    ws = wb.active
    blue_fill = PatternFill(fill_type="solid", fgColor="87CEEB")

    if vertical:
        # 原始列保持
        for col_idx, col_name in enumerate(df_a.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=1, column=len(df_a.columns)+1, value="匹配结果")

        current_row=2
        for i, gene in enumerate(df_a[gene_column_a]):
            info_row=df_a.iloc[i]
            matches=all_matches[i]
            flags=all_flags[i]

            if matches:
                for match, is_fuzzy in zip(matches, flags):
                    for col_idx, col_name in enumerate(df_a.columns,1):
                        ws.cell(row=current_row, column=col_idx, value=info_row[col_name])
                    cell = ws.cell(row=current_row, column=len(df_a.columns)+1, value=match)
                    if is_fuzzy:
                        cell.fill = blue_fill
                    current_row+=1
            else:
                # 保留空格行
                for col_idx, col_name in enumerate(df_a.columns,1):
                    ws.cell(row=current_row, column=col_idx, value=info_row[col_name])
                ws.cell(row=current_row, column=len(df_a.columns)+1, value="")
                current_row+=1
    else:
        # 横向排列
        max_len = max(len(m) for m in all_matches) if all_matches else 0
        ws.cell(row=1,column=1,value=gene_column_a)
        for j in range(max_len):
            ws.cell(row=1,column=2+j,value=f"匹配结果{j+1}")

        for i, gene in enumerate(df_a[gene_column_a],2):
            ws.cell(row=i,column=1,value=gene)
            matches=all_matches[i-2]
            flags=all_flags[i-2]
            for j in range(max_len):
                val = matches[j] if j<len(matches) else ""
                cell = ws.cell(row=i,column=2+j,value=val)
                if j<len(flags) and flags[j]:
                    cell.fill=blue_fill

    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)
    if sub_progress_callback:
        sub_progress_callback(100)

def gene_correspondence_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b,
                                      progress_callback=None, sub_progress_callback=None, set_progress_status=None):
    """基因查询（精确匹配横向）"""
    df_a=pd.read_excel(file_a)
    df_b=pd.read_excel(file_b)
    validate_columns(df_a,gene_column_a,set_progress_status=set_progress_status)
    validate_columns(df_b,gene_id_column_b,collinear_gene_column_b,set_progress_status=set_progress_status)
    a_to_b,b_to_a=build_bidirectional_map(df_b,gene_id_column_b,collinear_gene_column_b,set_progress_status=set_progress_status)

    results=[]
    total=len(df_a)
    for i,gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            results.append({gene_column_a:gene,"匹配结果":""})
            if progress_callback:
                progress_callback(int((i+1)/total*100))
            continue
        matches=a_to_b.get(gene,set())|b_to_a.get(gene,set())
        matches.discard(gene)
        row={gene_column_a:gene,"匹配结果":", ".join(map(str,matches)) if matches else ""}
        results.append(row)
        if progress_callback:
            progress_callback(int((i+1)/total*100))
    pd.DataFrame(results).to_excel(output_file,index=False,engine="openpyxl")
    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)
    if sub_progress_callback:
        sub_progress_callback(100)

# 精确匹配横向
def gene_search_with_progress(*args,**kwargs):
    return gene_correspondence_with_progress(*args,**kwargs)

# 精确匹配竖向
def classify_genes_with_progress(*args,**kwargs):
    df_a=pd.read_excel(args[0])
    df_b=pd.read_excel(args[1])
    gene_column_a=args[3]
    gene_id_column_b=args[4]
    collinear_gene_column_b=args[5]
    vertical=True
    return fuzzy_match_with_progress(args[0],args[1],args[2],
                                     gene_column_a,gene_id_column_b,collinear_gene_column_b,
                                     progress_callback=kwargs.get("progress_callback"),
                                     sub_progress_callback=kwargs.get("sub_progress_callback"),
                                     set_progress_status=kwargs.get("set_progress_status"),
                                     vertical=vertical)
