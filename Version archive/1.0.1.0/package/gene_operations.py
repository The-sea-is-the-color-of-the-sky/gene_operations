import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def classify_genes_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    """基因匹配功能，支持进度更新，竖向排列，A列和B列来回查找，重复3次"""
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    # 建立A->B和B->A映射
    a_to_b = {}
    b_to_a = {}
    for _, row in df_b.iterrows():
        a_val = row[gene_id_column_b]
        b_val = row[collinear_gene_column_b]
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    result_rows = []
    total = len(df_a)
    for i, (index, gene) in enumerate(df_a[gene_column_a].items()):
        if pd.isnull(gene):
            continue

        found = set()
        search_keys = set([gene])
        for _ in range(3):
            new_found = set()
            # A列查B列
            for key in search_keys:
                if key in a_to_b:
                    new_found.update(a_to_b[key])
            # B列查A列
            for key in search_keys:
                if key in b_to_a:
                    new_found.update(b_to_a[key])
            new_found -= found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found

        matches = list(found)
        if not matches:
            result_rows.append({gene_column_a: gene, "匹配结果": "无"})
        else:
            for match in matches:
                result_rows.append({gene_column_a: gene, "匹配结果": match})

        if progress_callback:
            progress_callback((i + 1) / total * 100)

    result_df = pd.DataFrame(result_rows)
    result_df.to_excel(output_file, index=False, engine="openpyxl")

def gene_correspondence_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    """基因对应功能，支持进度更新，遍历全表确保无遗漏"""
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    total = len(df_a)
    for i, (index, gene) in enumerate(df_a[gene_column_a].items()):
        if pd.isnull(gene):
            continue

        matches = []
        # 遍历整个信息表，分别判断A列和B列
        for _, row in df_b.iterrows():
            a_val = row[gene_id_column_b]
            b_val = row[collinear_gene_column_b]
            if a_val == gene and pd.notnull(b_val):
                matches.append(b_val)
            if b_val == gene and pd.notnull(a_val):
                matches.append(a_val)

        if not matches:
            df_a.at[index, "匹配结果"] = "无"
        else:
            df_a.at[index, "匹配结果"] = ", ".join(map(str, matches))

        if progress_callback:
            progress_callback((i + 1) / total * 100)

    df_a.to_excel(output_file, index=False, engine="openpyxl")

def gene_search_with_progress(fill_file, info_file, output_file, fill_col, info_a_col, info_b_col, update_progress=None, set_progress_status=None):
    from openpyxl import load_workbook
    if set_progress_status:
        set_progress_status("正在读取填入表格...")
    wb_fill = load_workbook(fill_file)
    ws_fill = wb_fill.active

    if set_progress_status:
        set_progress_status("正在读取信息表格...")
    wb_info = load_workbook(info_file)
    ws_info = wb_info.active

    if set_progress_status:
        set_progress_status("正在查找列索引...")
    # 获取列索引
    fill_col_idx = None
    for idx, cell in enumerate(ws_fill[1], 1):
        if cell.value == fill_col:
            fill_col_idx = idx
            break
    if fill_col_idx is None:
        raise ValueError(f"填入表格未找到列: {fill_col}")

    info_a_idx = None
    info_b_idx = None
    for idx, cell in enumerate(ws_info[1], 1):
        if cell.value == info_a_col:
            info_a_idx = idx
        if cell.value == info_b_col:
            info_b_idx = idx
    if info_a_idx is None or info_b_idx is None:
        raise ValueError("信息表格未找到A列或B列")

    if set_progress_status:
        set_progress_status("正在建立索引映射...")

    # 读取信息表所有数据，建立A列和B列的映射
    info_rows = list(ws_info.iter_rows(min_row=2, values_only=True))
    a_to_b = {}
    b_to_a = {}
    for row in info_rows:
        a_val = row[info_a_idx - 1]
        b_val = row[info_b_idx - 1]
        if a_val is not None and b_val is not None:
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    if set_progress_status:
        set_progress_status("正在统计最大匹配数...")

    # 匹配查找A列和B列之间来回查找，遍历整个表格，重复3次确保找到且没有遗漏
    all_matches = []
    max_match_count = 0
    rows = list(ws_fill.iter_rows(min_row=2))
    total = len(rows)
    for idx, row in enumerate(rows):
        key = row[fill_col_idx - 1].value
        found = set()
        search_keys = set([key])
        for _ in range(3):
            new_found = set()
            # 遍历整个信息表，A列查B列
            for a_val, b_val in zip(
                [r[info_a_idx - 1] for r in info_rows],
                [r[info_b_idx - 1] for r in info_rows]
            ):
                if a_val in search_keys and b_val is not None:
                    new_found.add(b_val)
            # 遍历整个信息表，B列查A列
            for a_val, b_val in zip(
                [r[info_a_idx - 1] for r in info_rows],
                [r[info_b_idx - 1] for r in info_rows]
            ):
                if b_val in search_keys and a_val is not None:
                    new_found.add(a_val)
            new_found -= found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found
        matches = list(found)
        all_matches.append(matches)
        if len(matches) > max_match_count:
            max_match_count = len(matches)
        if update_progress:
            update_progress(int((idx + 1) / total * 50))
        if set_progress_status and idx == 0:
            set_progress_status("正在查找匹配信息...")

    if set_progress_status:
        set_progress_status("正在写入表头...")
    # 添加表头
    start_col = ws_fill.max_column + 1
    for j in range(max_match_count):
        ws_fill.cell(row=1, column=start_col + j, value=f"匹配结果{j}")

    if set_progress_status:
        set_progress_status("正在写入匹配结果...")
    # 填充数据
    for i, matches in enumerate(all_matches, 2):
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws_fill.cell(row=i, column=start_col + j, value=val)
        if update_progress:
            update_progress(50 + int((i - 1) / len(rows) * 50))

    wb_fill.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    if update_progress:
        update_progress(100)

def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None, set_progress_status=None):
    """
    模糊匹配，完全匹配和原有逻辑不变，仅对不完全匹配部分用蓝色高亮，不添加匹配类型列，支持横向多结果。
    横向排列：输出和gene_search_with_progress一致，只是多了模糊匹配的结果和高亮。
    """
    import pandas as pd
    from openpyxl import Workbook

    if set_progress_status:
        set_progress_status("正在读取表格...")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    # 先收集所有匹配结果，确定最大匹配数
    all_matches = []
    all_fuzzy_flags = []
    max_match_count = 0
    total = len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_fuzzy_flags.append([])
            continue
        matches = []
        match_is_fuzzy = []
        # 完全匹配
        exact_matches = df_b[df_b[gene_id_column_b] == gene][collinear_gene_column_b].tolist()
        exact_matches += df_b[df_b[collinear_gene_column_b] == gene][gene_id_column_b].tolist()
        for m in exact_matches:
            matches.append(m)
            match_is_fuzzy.append(False)
        # 模糊匹配
        fuzzy_matches = []
        for col in [gene_id_column_b, collinear_gene_column_b]:
            fuzzy = df_b[df_b[col].astype(str).str.contains(str(gene), na=False) & (df_b[col] != gene)][col].tolist()
            fuzzy_matches += fuzzy
        fuzzy_matches = [m for m in set(fuzzy_matches) if m not in matches]
        for m in fuzzy_matches:
            matches.append(m)
            match_is_fuzzy.append(True)
        if not matches:
            matches = ["无"]
            match_is_fuzzy = [False]
        all_matches.append(matches)
        all_fuzzy_flags.append(match_is_fuzzy)
        if len(matches) > max_match_count:
            max_match_count = len(matches)
        if progress_callback:
            progress_callback((i + 1) / total * 100)

    if set_progress_status:
        set_progress_status("正在写入Excel...")

    # 写入Excel，横向多结果
    wb = Workbook()
    ws = wb.active
    # 写表头
    ws.cell(row=1, column=1, value=gene_column_a)
    for j in range(max_match_count):
        ws.cell(row=1, column=2 + j, value=f"匹配结果{j+1}")
    highlight_cells = []
    for i, gene in enumerate(df_a[gene_column_a]):
        ws.cell(row=i + 2, column=1, value=gene)
        matches = all_matches[i]
        fuzzy_flags = all_fuzzy_flags[i]
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws.cell(row=i + 2, column=2 + j, value=val)
            if j < len(fuzzy_flags) and fuzzy_flags[j]:
                col_letter = ws.cell(row=1, column=2 + j).column_letter
                highlight_cells.append(f"{col_letter}{i+2}")
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    return highlight_cells, output_file

def fuzzy_match_with_progress_vertical(
    file_a, file_b, output_file,
    gene_column_a, gene_id_column_b, collinear_gene_column_b,
    progress_callback=None, set_progress_status=None
):
    """
    竖向模糊匹配，A列和B列来回查找，重复3次，完全匹配和原有逻辑不变，仅对不完全匹配部分用蓝色高亮，不添加匹配类型列，支持竖向多结果。
    """
    import pandas as pd
    from openpyxl import Workbook

    if set_progress_status:
        set_progress_status("正在读取表格...")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    # 建立A->B和B->A映射
    a_to_b = {}
    b_to_a = {}
    for _, row in df_b.iterrows():
        a_val = row[gene_id_column_b]
        b_val = row[collinear_gene_column_b]
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    result_rows = []
    highlight_cells = []
    total = len(df_a)
    excel_row = 2  # Excel行号，从2开始（1为表头）

    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        found = set()
        found_type = dict()
        search_keys = set([gene])
        for round_idx in range(3):
            if set_progress_status:
                set_progress_status(f"第{round_idx+1}/3轮查找：正在处理第 {i+1}/{total} 个基因...")
            new_found = set()
            new_type = dict()
            # A列查B列
            for key in search_keys:
                if key in a_to_b:
                    for b_val in a_to_b[key]:
                        if b_val not in found:
                            new_found.add(b_val)
                            new_type[b_val] = False
            # B列查A列
            for key in search_keys:
                if key in b_to_a:
                    for a_val in b_to_a[key]:
                        if a_val not in found:
                            new_found.add(a_val)
                            new_type[a_val] = False
            # 模糊匹配
            for key in search_keys:
                # A列模糊查B列
                for a_val in a_to_b:
                    if pd.notnull(a_val) and str(key) in str(a_val) and a_val != key:
                        for b_val in a_to_b[a_val]:
                            if b_val not in found:
                                new_found.add(b_val)
                                new_type[b_val] = True
                # B列模糊查A列
                for b_val in b_to_a:
                    if pd.notnull(b_val) and str(key) in str(b_val) and b_val != key:
                        for a_val in b_to_a[b_val]:
                            if a_val not in found:
                                new_found.add(a_val)
                                new_type[a_val] = True
            new_found -= found
            if not new_found:
                break
            for m in new_found:
                found_type[m] = new_type[m]
            found.update(new_found)
            search_keys = new_found
        matches = list(found)
        if not matches:
            result_rows.append([gene, "无"])
            excel_row += 1
        else:
            for m in matches:
                result_rows.append([gene, m])
                if found_type.get(m, False):
                    highlight_cells.append(f"B{excel_row}")
                excel_row += 1
        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"正在处理第 {i+1}/{total} 个基因...")

    if set_progress_status:
        set_progress_status("正在写入Excel...")

    wb = Workbook()
    ws = wb.active
    ws.append([gene_column_a, "匹配结果"])
    for row in result_rows:
        ws.append(row)
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    return highlight_cells, output_file
    for row in result_rows:
        ws.append(row)
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    return highlight_cells, output_file
