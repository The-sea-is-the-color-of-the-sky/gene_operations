import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

from package.gene_operations import (
    classify_genes_with_progress,
    gene_search_with_progress,
    fuzzy_match_with_progress,
    gene_correspondence_with_progress
)

selected_function = None
selected_match_mode = None

# 在全局提前声明 fuzzy_var
fuzzy_var = None

# 在全局提前声明控件变量
entry_file_a = None
entry_file_b = None
entry_output_file = None
entry_gene_column_a = None
entry_gene_id_column_b = None
entry_collinear_gene_column_b = None
progress_bar = None
progress_status_label = None

# ---------------- 文件选择与列名更新 ----------------
def select_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def select_output_file(entry):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def update_column_options(entry, file_path):
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        entry["values"] = columns
    except Exception as e:
        messagebox.showerror("错误", f"无法读取列名：{str(e)}")

def select_file_and_update_columns(entry_file, *entry_columns):
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)
        for entry_column in entry_columns:
            update_column_options(entry_column, file_path)

# ---------------- 核心功能运行函数 ----------------
def run_function():
    if selected_function == "基因查询":
        run_gene_correspondence()
    elif selected_function == "基因匹配":
        is_fuzzy = fuzzy_var.get()
        if selected_match_mode == "横向排列":
            run_fuzzy_match(horizontal=True) if is_fuzzy else run_gene_search()
        elif selected_match_mode == "竖向排列":
            run_fuzzy_match(horizontal=False) if is_fuzzy else run_classification()
        else:
            messagebox.showerror("错误", "请选择基因匹配的排列方式！")
    else:
        messagebox.showerror("错误", "请选择功能类型！")

def run_gene_correspondence():
    run_threaded(gene_correspondence_with_progress)

def run_gene_search():
    run_threaded(gene_search_with_progress)

def run_classification():
    run_threaded(classify_genes_with_progress, classification=True)

def run_fuzzy_match(horizontal=True):
    if horizontal:
        run_threaded(fuzzy_match_with_progress)
    else:
        run_threaded(lambda *args, **kwargs: fuzzy_match_with_progress(*args, **kwargs, vertical=True), special=True)

# ---------------- 通用运行封装 ----------------
def get_common_args():
    return (
        entry_file_a.get(),
        entry_file_b.get(),
        entry_output_file.get(),
        entry_gene_column_a.get(),
        entry_gene_id_column_b.get(),
        entry_collinear_gene_column_b.get(),
    )

def run_threaded(func, classification=False, special=False):
    args = get_common_args()
    if not all(args):
        messagebox.showerror("错误", "请填写所有选项！")
        return

    def task():
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            if classification:
                func(*args, progress_callback=set_progress)
            elif special:
                result = func(*args, progress_callback=set_progress)
                if isinstance(result, tuple) and len(result) == 2:
                    highlight_cells, output_file = result
                    if highlight_cells:
                        highlight_cells_in_excel(output_file, highlight_cells)
            else:
                func(*args, progress_callback=set_progress, set_progress_status=set_progress_status)
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

def highlight_cells_in_excel(file_path, highlight_cells):
    wb = load_workbook(file_path)
    ws = wb.active
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in highlight_cells:
        ws[cell].fill = blue_fill
    wb.save(file_path)

# ---------------- 进度条 ----------------
def set_progress(value):
    root.after(0, lambda: progress_bar.config(value=value))

def set_progress_status(text):
    root.after(0, lambda: progress_status_label.config(text=text))

def reset_progress():
    set_progress(0)
    set_progress_status("等待运行...")

# ---------------- GUI 初始化 ----------------
def main():
    global root
    global fuzzy_var
    global entry_file_a, entry_file_b, entry_output_file
    global entry_gene_column_a, entry_gene_id_column_b, entry_collinear_gene_column_b
    global progress_bar, progress_status_label
    root = tk.Tk()
    root.title("基因工具")
    root.geometry("650x550")
    root.resizable(False, False)

    # 修复icon.ico路径问题，若不存在则跳过设置
    icon_path = os.path.join(os.path.dirname(__file__), "package", "icon.ico")
    if os.path.exists(icon_path):
        try:
            root.iconbitmap(icon_path)
        except Exception:
            pass

    main_frame = ttk.Notebook(root)
    main_frame.pack(fill="both", expand=True)

    frame_function = ttk.LabelFrame(main_frame, text="选择功能", padding="10")
    frame_files = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
    frame_columns = ttk.LabelFrame(main_frame, text="列名输入", padding="10")
    frame_progress = ttk.LabelFrame(main_frame, text="运行进度", padding="10")

    for f in [frame_function, frame_files, frame_columns, frame_progress]:
        f.pack(fill="x", pady=5)

    # 功能区控件
    ttk.Label(frame_function, text="功能类型:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    function_combo = ttk.Combobox(frame_function, values=["基因查询", "基因匹配"], state="readonly", width=20)
    function_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    match_mode_label = ttk.Label(frame_function, text="排列方式:")
    match_mode_combo = ttk.Combobox(frame_function, values=["横向排列", "竖向排列"], state="readonly", width=18)
    fuzzy_var = tk.BooleanVar(value=False)  # 这里赋值
    fuzzy_check = ttk.Checkbutton(frame_function, text="启用模糊匹配", variable=fuzzy_var)

    match_mode_label.grid_remove()
    match_mode_combo.grid_remove()
    fuzzy_check.grid_remove()

    def on_function_select(event):
        global selected_function, selected_match_mode
        selected_function = function_combo.get()
        if selected_function == "基因匹配":
            match_mode_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")
            match_mode_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
            fuzzy_check.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        else:
            match_mode_label.grid_remove()
            match_mode_combo.grid_remove()
            fuzzy_check.grid_remove()
            selected_match_mode = None

    def on_match_mode_select(event):
        global selected_match_mode
        selected_match_mode = match_mode_combo.get()

    function_combo.bind("<<ComboboxSelected>>", on_function_select)
    match_mode_combo.bind("<<ComboboxSelected>>", on_match_mode_select)

    # 文件选择区
    ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    entry_file_a = ttk.Entry(frame_files, width=50)
    entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
    ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5)

    ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    entry_file_b = ttk.Entry(frame_files, width=50)
    entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
    ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5)

    ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
    entry_output_file = ttk.Entry(frame_files, width=50)
    entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
    ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5)

    # 列名输入区
    ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")
    entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
    entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
    entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
    entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # 运行与进度
    progress_bar = ttk.Progressbar(frame_progress, orient="horizontal", length=400, mode="determinate", maximum=100)
    progress_bar.pack(fill="x", padx=5, pady=5)
    progress_status_label = ttk.Label(frame_progress, text="等待运行...", anchor="w")
    progress_status_label.pack(fill="x", padx=5, pady=(0, 5))

    run_button = ttk.Button(frame_progress, text="运行", command=run_function)
    run_button.pack(pady=10)

    # 启动主循环
    root.mainloop()

if __name__ == "__main__":
    main()
