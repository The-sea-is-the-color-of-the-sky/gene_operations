import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# ======= 通用工具函数 =======
# 这些函数为后续基因操作提供通用的数据校验、映射关系构建和递归搜索等基础功能。

def validate_columns(df, *columns, set_progress_status=None):
    """验证DataFrame中是否包含指定的列名。"""
    if set_progress_status:
        set_progress_status("验证列名是否存在")
    for col in columns:
        if col not in df.columns:
            raise ValueError(f"表格中不存在列名 '{col}'")

def build_bidirectional_map(df, col1, col2, set_progress_status=None):
    """根据两列内容构建双向映射关系（字典）。"""
    if set_progress_status:
        set_progress_status("构建双向映射表")
    a_to_b, b_to_a = {}, {}
    for a_val, b_val in zip(df[col1], df[col2]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)
    return a_to_b, b_to_a

def recursive_search(keys, a_to_b, b_to_a, depth=3, fuzzy=False, set_progress_status=None):
    """
    递归地在双向映射关系中查找与keys相关的所有基因。
    支持精确和模糊两种查找方式。
    """
    if set_progress_status:
        set_progress_status("递归搜索匹配基因")
    found = set(keys)
    current = set(keys)
    for _ in range(depth):
        next_found = set()
        for key in current:
            if fuzzy:
                next_found.update(b for a in a_to_b if key in str(a) and a != key for b in a_to_b[a])
                next_found.update(a for b in b_to_a if key in str(b) and b != key for a in b_to_a[b])
            else:
                next_found |= a_to_b.get(key, set())
                next_found |= b_to_a.get(key, set())
        new = next_found - found
        if not new:
            break
        found |= new
        current = new
    return found - set(keys)

# ======= 核心功能函数（附加状态标签） =======
# 以下为主要的基因操作功能函数，均支持进度回调和状态更新。

def classify_genes_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, 
                                progress_callback=None, set_progress_status=None):
    """
    对输入文件A中的基因，在文件B中递归查找其所有相关基因，并输出匹配结果。
    支持进度和状态回调。
    """
    if set_progress_status:
        set_progress_status("读取输入文件")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    validate_columns(df_a, gene_column_a, set_progress_status=set_progress_status)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)

    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)

    if set_progress_status:
        set_progress_status("开始基因匹配")
    results = []
    total = len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        found = recursive_search({gene}, a_to_b, b_to_a, depth=3, set_progress_status=set_progress_status)
        for match in found:
            results.append({gene_column_a: gene, "匹配结果": match})
        if progress_callback:
            progress_callback((i + 1) / total * 100)

    if set_progress_status:
        set_progress_status("保存匹配结果")
    pd.DataFrame(results).to_excel(output_file, index=False, engine="openpyxl")
    if set_progress_status:
        set_progress_status("保存完成")


def gene_search_with_progress(fill_file, info_file, output_file, fill_col, info_a_col, info_b_col, 
                              progress_callback=None, set_progress_status=None):
    """
    在填入表中，根据信息表递归查找每个基因的所有相关基因，并将结果写入新表。
    支持进度和状态回调。
    """
    if set_progress_status:
        set_progress_status("读取填入表...")
    wb_fill = load_workbook(fill_file)
    ws_fill = wb_fill.active

    if set_progress_status:
        set_progress_status("读取信息表...")
    wb_info = load_workbook(info_file)
    ws_info = wb_info.active

    if set_progress_status:
        set_progress_status("提取列索引")
    fill_idx = next((idx for idx, cell in enumerate(ws_fill[1], 1) if cell.value == fill_col), None)
    info_a_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_a_col), None)
    info_b_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_b_col), None)

    if None in (fill_idx, info_a_idx, info_b_idx):
        raise ValueError("缺少必要列")

    if set_progress_status:
        set_progress_status("构建映射关系")
    info_rows = list(ws_info.iter_rows(min_row=2, values_only=True))
    a_to_b, b_to_a = {}, {}
    for row in info_rows:
        a_val = row[info_a_idx - 1]
        b_val = row[info_b_idx - 1]
        if a_val is not None and b_val is not None:
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    if set_progress_status:
        set_progress_status("执行匹配计算")
    rows = list(ws_fill.iter_rows(min_row=2))
    total = len(rows)
    max_match_count = 0
    all_matches = []

    for idx, row in enumerate(rows):
        key = row[fill_idx - 1].value
        found = recursive_search({key}, a_to_b, b_to_a, depth=3) if key else set()
        matches = list(found)
        all_matches.append(matches)
        max_match_count = max(max_match_count, len(matches))
        if progress_callback:
            progress_callback(int((idx + 1) / total * 50))

    if set_progress_status:
        set_progress_status("写入匹配结果")
    start_col = ws_fill.max_column + 1
    for j in range(max_match_count):
        ws_fill.cell(row=1, column=start_col + j, value=f"匹配结果{j+1}")

    for i, matches in enumerate(all_matches, 2):
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws_fill.cell(row=i, column=start_col + j, value=val)
        if progress_callback:
            progress_callback(50 + int((i - 1) / total * 50))

    if set_progress_status:
        set_progress_status("保存结果文件")
    wb_fill.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)


def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, 
                              progress_callback=None, set_progress_status=None, vertical=False):
    """
    对输入文件A中的基因，分别进行精确和模糊递归匹配，并将结果写入新Excel。
    模糊匹配结果以蓝色高亮显示。
    vertical: 若为True，结果竖向排列，每个匹配结果单独占一行，原始信息列保持
    """

 # 读取文件
    if set_progress_status:
        set_progress_status("读取文件")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    # 验证列
    validate_columns(df_a, gene_column_a, set_progress_status=set_progress_status)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)
    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)

    # 匹配计算
    all_matches = []
    all_flags = []
    total = len(df_a)

    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_flags.append([])
            continue

        exact = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=False, set_progress_status=set_progress_status)
        fuzzy = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=True, set_progress_status=set_progress_status)

        exact_list = list(exact)
        fuzzy_list = list(fuzzy - exact)

        matches = exact_list + fuzzy_list
        flags = [False] * len(exact_list) + [True] * len(fuzzy_list)

        all_matches.append(matches)
        all_flags.append(flags)

        if progress_callback:
            progress_callback(int((i+1)/total*100))
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"处理进度：{i+1}/{total}")

    # 写入 Excel
    wb = Workbook()
    ws = wb.active
    blue_fill = PatternFill(fill_type="solid", fgColor="87CEEB")

    if vertical:
        # 竖向排列
        for col_idx, col_name in enumerate(df_a.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=1, column=len(df_a.columns)+1, value="匹配结果")

        current_row = 2
        for i, gene in enumerate(df_a[gene_column_a]):
            info_row = df_a.iloc[i]
            matches = all_matches[i]
            flags = all_flags[i]

            if matches:
                for match, is_fuzzy in zip(matches, flags):
                    # 复制原始信息列
                    for col_idx, col_name in enumerate(df_a.columns, 1):
                        ws.cell(row=current_row, column=col_idx, value=info_row[col_name])
                    # 写匹配结果
                    cell = ws.cell(row=current_row, column=len(df_a.columns)+1, value=match)
                    if is_fuzzy:
                        cell.fill = blue_fill
                    current_row += 1
                current_row += 1  # 每个基因匹配完空一行
            else:
                # 无匹配结果保留原始信息
                for col_idx, col_name in enumerate(df_a.columns, 1):
                    ws.cell(row=current_row, column=col_idx, value=info_row[col_name])
                current_row += 2
    else:
        # 横向排列
        max_match_count = max(len(m) for m in all_matches) if all_matches else 0
        ws.cell(row=1, column=1, value=gene_column_a)
        for j in range(max_match_count):
            ws.cell(row=1, column=2+j, value=f"匹配结果{j+1}")

        for i, gene in enumerate(df_a[gene_column_a], 2):
            ws.cell(row=i, column=1, value=gene)
            matches = all_matches[i-2]
            flags = all_flags[i-2]
            for j in range(max_match_count):
                val = matches[j] if j < len(matches) else ""
                cell = ws.cell(row=i, column=2+j, value=val)
                if j < len(flags) and flags[j]:
                    cell.fill = blue_fill

    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")

    return [], output_file



def gene_correspondence_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, 
                                      progress_callback=None, set_progress_status=None):
    """
    查找输入文件A中每个基因在文件B中的直接对应关系，并输出结果。
    支持进度和状态回调。
    """
    if set_progress_status:
        set_progress_status("读取Excel文件")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    validate_columns(df_a, gene_column_a, set_progress_status=set_progress_status)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)
    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b, set_progress_status=set_progress_status)

    total = len(df_a)
    result_rows = []

    if set_progress_status:
        set_progress_status("查找对应关系")
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            row = df_a.iloc[i].copy()
            row["匹配结果"] = ""
            result_rows.append(row)
            if progress_callback:
                progress_callback((i + 1) / total * 100)
            continue

        matches = a_to_b.get(gene, set()) | b_to_a.get(gene, set())
        matches.discard(gene)
        row = df_a.iloc[i].copy()
        if matches:
            row["匹配结果"] = ", ".join(map(str, matches))
        else:
            row["匹配结果"] = ""
        result_rows.append(row)

        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"处理进度：{i + 1}/{total}")

    result_df = pd.DataFrame(result_rows)
    result_df.to_excel(output_file, index=False, engine="openpyxl")

    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)
