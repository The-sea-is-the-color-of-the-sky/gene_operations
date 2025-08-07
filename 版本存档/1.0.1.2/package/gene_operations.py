import pandas as pd  # 导入pandas用于数据处理
from openpyxl import load_workbook  # 导入openpyxl用于Excel读写
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def classify_genes_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    """基因匹配功能，支持进度更新，竖向排列，A列和B列来回查找，重复3次"""
    df_a = pd.read_excel(file_a)  # 读取A表
    df_b = pd.read_excel(file_b)  # 读取B表

    # 检查列名是否存在
    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    # 优化映射构建
    a_to_b = {}  # A到B的映射
    b_to_a = {}  # B到A的映射
    for a_val, b_val in zip(df_b[gene_id_column_b], df_b[collinear_gene_column_b]):
        if pd.notnull(a_val) and pd.notnull(b_val):  # 只处理非空值
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    result_rows = []  # 存储结果
    total = len(df_a)  # 总行数
    for i, gene in enumerate(df_a[gene_column_a]):  # 遍历A表每个基因
        if pd.isnull(gene):  # 跳过空值
            continue
        found = set()  # 已找到的匹配集合
        search_keys = {gene}  # 当前查找的key集合
        for _ in range(3):  # 来回查找3轮
            # 用集合推导优化查找
            new_found = set().union(
                *(a_to_b.get(key, set()) for key in search_keys),
                *(b_to_a.get(key, set()) for key in search_keys)
            ) - found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found
        if found:
            result_rows.extend([{gene_column_a: gene, "匹配结果": match} for match in found])
        if progress_callback:
            progress_callback((i + 1) / total * 100)  # 进度回调
    pd.DataFrame(result_rows).to_excel(output_file, index=False, engine="openpyxl")  # 写入Excel

def gene_search_with_progress(fill_file, info_file, output_file, fill_col, info_a_col, info_b_col, progress_callback=None, set_progress_status=None):
    from openpyxl import load_workbook
    if set_progress_status:
        set_progress_status("正在读取填入表格...")  # 状态提示
    wb_fill = load_workbook(fill_file)  # 读取填入表
    ws_fill = wb_fill.active

    if set_progress_status:
        set_progress_status("正在读取信息表格...")  # 状态提示
    wb_info = load_workbook(info_file)  # 读取信息表
    ws_info = wb_info.active

    if set_progress_status:
        set_progress_status("正在查找列索引...")  # 状态提示
    fill_col_idx = next((idx for idx, cell in enumerate(ws_fill[1], 1) if cell.value == fill_col), None)  # 查找填入列索引
    if fill_col_idx is None:
        raise ValueError(f"填入表格未找到列: {fill_col}")

    info_a_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_a_col), None)  # 查找A列索引
    info_b_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_b_col), None)  # 查找B列索引
    if info_a_idx is None or info_b_idx is None:
        raise ValueError("信息表格未找到A列或B列")

    if set_progress_status:
        set_progress_status("正在建立索引映射...")  # 状态提示

    info_rows = list(ws_info.iter_rows(min_row=2, values_only=True))  # 读取信息表所有数据行
    a_to_b = {}  # A到B映射
    b_to_a = {}  # B到A映射
    for row in info_rows:
        a_val = row[info_a_idx - 1]
        b_val = row[info_b_idx - 1]
        if a_val is not None and b_val is not None:
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    if set_progress_status:
        set_progress_status("正在统计最大匹配数...")  # 状态提示

    rows = list(ws_fill.iter_rows(min_row=2))  # 读取填入表所有数据行
    total = len(rows)  # 总行数
    all_matches = []  # 所有匹配结果
    max_match_count = 0  # 最大匹配数
    for idx, row in enumerate(rows):
        key = row[fill_col_idx - 1].value  # 获取当前行的key
        found = set()
        search_keys = {key}
        for _ in range(3):  # 来回查找3轮
            new_found = set().union(
                *(a_to_b.get(k, set()) for k in search_keys),
                *(b_to_a.get(k, set()) for k in search_keys)
            ) - found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found
        matches = list(found)
        if matches:
            all_matches.append(matches)
            max_match_count = max(max_match_count, len(matches))
        else:
            all_matches.append([])
        if progress_callback:
            progress_callback(int((idx + 1) / total * 50))  # 进度回调
        if set_progress_status and idx == 0:
            set_progress_status("正在查找匹配信息...")  # 状态提示

    if set_progress_status:
        set_progress_status("正在写入表头...")  # 状态提示
    start_col = ws_fill.max_column + 1  # 新增列起始位置
    for j in range(max_match_count):
        ws_fill.cell(row=1, column=start_col + j, value=f"匹配结果{j+1}")  # 写入表头

    if set_progress_status:
        set_progress_status("正在写入匹配结果...")  # 状态提示
    for i, matches in enumerate(all_matches, 2):
        if not matches:
            continue
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws_fill.cell(row=i, column=start_col + j, value=val)  # 写入匹配结果
        if progress_callback:
            progress_callback(50 + int((i - 1) / len(rows) * 50))  # 进度回调
    wb_fill.save(output_file)  # 保存Excel
    if set_progress_status:
        set_progress_status("保存完成")  # 状态提示
    if progress_callback:
        progress_callback(100)  # 进度回调

def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None, set_progress_status=None):
    """
    横向排列：遍历全表多次查找，完全匹配和模糊匹配都支持多结果，输出多列，模糊匹配的单元格标为蓝色。
    匹配不到时不填无，直接跳过该行。
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    if set_progress_status:
        set_progress_status("正在读取填入表格...")
    df_a = pd.read_excel(file_a)
    if set_progress_status:
        set_progress_status("正在读取信息表格...")
    df_b = pd.read_excel(file_b)

    # 检查列名是否存在
    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    if set_progress_status:
        set_progress_status("正在建立映射...")

    a_to_b = {}
    b_to_a = {}
    for a_val, b_val in zip(df_b[gene_id_column_b], df_b[collinear_gene_column_b]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    # 递归查找完全匹配
    def recursive_exact(keys, a_to_b, b_to_a, found, depth):
        if depth == 0:
            return set()
        new_found = set().union(
            *(a_to_b.get(k, set()) for k in keys),
            *(b_to_a.get(k, set()) for k in keys)
        ) - found
        if not new_found:
            return set()
        found.update(new_found)
        return new_found | recursive_exact(new_found, a_to_b, b_to_a, found, depth - 1)

    # 递归查找模糊匹配
    def recursive_fuzzy(gene, a_to_b, b_to_a, found, depth):
        if depth == 0:
            return set()
        fuzzy_a = {b for a in a_to_b if pd.notnull(a) and str(gene) in str(a) and a != gene for b in a_to_b[a]}
        fuzzy_b = {a for b in b_to_a if pd.notnull(b) and str(gene) in str(b) and b != gene for a in b_to_a[b]}
        new_found = (fuzzy_a | fuzzy_b) - found
        if not new_found:
            return set()
        found.update(new_found)
        return new_found | recursive_fuzzy(gene, a_to_b, b_to_a, found, depth - 1)

    all_matches = []
    all_fuzzy_flags = []
    max_match_count = 0
    total = len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_fuzzy_flags.append([])
            continue
        # 完全匹配
        found_exact = set()
        exact_matches = recursive_exact({gene}, a_to_b, b_to_a, found_exact, 3)
        found = list(exact_matches)
        found_flags = [False] * len(found)
        # 模糊匹配
        found_fuzzy = set(found)
        fuzzy_matches = recursive_fuzzy(gene, a_to_b, b_to_a, found_fuzzy, 3)
        found += list(fuzzy_matches)
        found_flags += [True] * len(fuzzy_matches)
        if found:
            all_matches.append(found)
            all_fuzzy_flags.append(found_flags)
            max_match_count = max(max_match_count, len(found))
        else:
            all_matches.append([])
            all_fuzzy_flags.append([])
        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"正在处理第 {i+1}/{total} 个基因...")

    if set_progress_status:
        set_progress_status("正在写入Excel...")

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=gene_column_a)
    for j in range(max_match_count):
        ws.cell(row=1, column=2 + j, value=f"匹配结果{j+1}")
    blue_fill = PatternFill(fill_type="solid", fgColor="87CEEB")
    excel_row = 2
    for i, gene in enumerate(df_a[gene_column_a]):
        matches = all_matches[i]
        fuzzy_flags = all_fuzzy_flags[i]
        # 保留所有行，无论是否有匹配
        ws.cell(row=excel_row, column=1, value=gene)
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            cell = ws.cell(row=excel_row, column=2 + j, value=val)
            if j < len(fuzzy_flags) and fuzzy_flags[j]:
                cell.fill = blue_fill  # 模糊匹配标蓝色
        excel_row += 1
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    # 返回所有被高亮的单元格（蓝色填充的单元格坐标），便于外部调用者进一步处理
    highlight_cells = []
    excel_row = 2
    for i, matches in enumerate(all_matches):
        fuzzy_flags = all_fuzzy_flags[i]
        if not matches:
            continue
        for j in range(len(matches)):
            if j < len(fuzzy_flags) and fuzzy_flags[j]:
                # 匹配结果列从2开始
                col_letter = ws.cell(row=1, column=2 + j).column_letter
                highlight_cells.append(f"{col_letter}{excel_row}")
        excel_row += 1
    return highlight_cells, output_file  # 返回高亮单元格和文件路径

def gene_correspondence_with_progress(
    file_a, file_b, output_file,highlight_cells,
    gene_column_a, gene_id_column_b, collinear_gene_column_b,
    progress_callback=None, set_progress_status=None
):
    """
    基因查询功能，支持进度与状态回调。
    file_a: 查询表格路径
    file_b: 信息表格路径
    output_file: 输出文件路径
    gene_column_a: 查询表格目标列名
    gene_id_column_b: 信息表A列名
    collinear_gene_column_b: 信息表B列名
    progress_callback: 进度回调(百分比)
    set_progress_status: 状态文本回调
    """
    import pandas as pd

    # 状态提示：读取表格
    if set_progress_status:
        set_progress_status("正在读取填入表格...")
    df_a = pd.read_excel(file_a)
    if set_progress_status:
        set_progress_status("正在读取信息表格...")
    df_b = pd.read_excel(file_b)

    # 检查列名
    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    total = len(df_a)
    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    # 预处理B表为查找字典，加速匹配
    a_to_b = {}
    b_to_a = {}
    for _, row in df_b.iterrows():
        a_val = row[gene_id_column_b]
        b_val = row[collinear_gene_column_b]
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    # 匹配主循环
    match_results = []
    for i, (index, gene) in enumerate(df_a[gene_column_a].items()):
        if pd.isnull(gene):
            match_results.append("无")
        else:
            matches = set()
            # A->B
            if gene in a_to_b:
                matches.update(a_to_b[gene])
            # B->A
            if gene in b_to_a:
                matches.update(b_to_a[gene])
            # 去除自身
            matches.discard(gene)
            if matches:
                match_results.append(", ".join(map(str, matches)))
            else:
                match_results.append("无")
        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"正在处理第 {i+1}/{total} 个基因...")

    # 写入结果
    df_a["匹配结果"] = match_results
    if set_progress_status:
        set_progress_status("正在写入Excel...")
    df_a.to_excel(output_file, index=False, engine="openpyxl")
    if set_progress_status:
        set_progress_status("操作完成")
    if progress_callback:
        progress_callback(100)
        set_progress_status("保存完成")  # 状态提示
    return highlight_cells, output_file  # 返回高亮单元格和文件路径

def gene_correspondence_with_progress(
    file_a, file_b, output_file,
    gene_column_a, gene_id_column_b, collinear_gene_column_b,
    progress_callback=None, set_progress_status=None
):
    """
    基因查询功能，支持进度与状态回调。
    file_a: 查询表格路径
    file_b: 信息表格路径
    output_file: 输出文件路径
    gene_column_a: 查询表格目标列名
    gene_id_column_b: 信息表A列名
    collinear_gene_column_b: 信息表B列名
    progress_callback: 进度回调(百分比)
    set_progress_status: 状态文本回调
    """
    import pandas as pd

    if set_progress_status:
        set_progress_status("正在读取填入表格...")
    df_a = pd.read_excel(file_a)
    if set_progress_status:
        set_progress_status("正在读取信息表格...")
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    total = len(df_a)
    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    for i, (index, gene) in enumerate(df_a[gene_column_a].items()):
        if pd.isnull(gene):
            df_a.at[index, "匹配结果"] = "无"
            if progress_callback:
                progress_callback((i + 1) / total * 100)
            continue
        matches = []
        for _, row in df_b.iterrows():
            a_val = row[gene_id_column_b]
            b_val = row[collinear_gene_column_b]
            if a_val == gene and pd.notnull(b_val):
                matches.append(b_val)
            if b_val == gene and pd.notnull(a_val):
                matches.append(a_val)
        if not matches:
            df_a.at[index, "匹配结果"] = "无"
        else:
            df_a.at[index, "匹配结果"] = ", ".join(map(str, matches))
        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"正在处理第 {i+1}/{total} 个基因...")

    if set_progress_status:
        set_progress_status("正在写入Excel...")
    df_a.to_excel(output_file, index=False, engine="openpyxl")
    if set_progress_status:
        set_progress_status("操作完成")
    if progress_callback:
        progress_callback(100)



