import tkinter as tk
from tkinter import ttk, filedialog, messagebox, LabelFrame
from package.gene_operations import (
    classify_genes_with_progress,
    gene_correspondence_with_progress,
    gene_search_with_progress,
    fuzzy_match_with_progress,  # 新增导入
    fuzzy_match_with_progress_vertical  # 新增竖向模糊匹配导入
)
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# 全局变量
selected_function = None
selected_match_mode = None  # 新增变量

def select_file(entry):
    """打开文件选择对话框并设置文件路径"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def select_output_file(entry):
    """选择输出文件路径"""
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def update_column_options(entry, file_path):
    """更新列名下拉选项"""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        entry["values"] = columns
    except Exception as e:
        messagebox.showerror("错误", f"无法读取列名：{str(e)}")

def select_file_and_update_columns(entry_file, *entry_columns):
    """选择文件并更新多个列名选项"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)
        for entry_column in entry_columns:
            update_column_options(entry_column, file_path)

def run_function():
    """运行所选功能"""
    if selected_function == "基因查询":
        run_gene_correspondence()
    elif selected_function == "基因匹配":
        is_fuzzy = fuzzy_var.get()
        if selected_match_mode == "横向排列":
            if is_fuzzy:
                run_fuzzy_match(horizontal=True)
            else:
                run_gene_search()
        elif selected_match_mode == "竖向排列":
            if is_fuzzy:
                run_fuzzy_match(horizontal=False)
            else:
                run_classification()
        else:
            messagebox.showerror("错误", "请选择基因匹配的排列方式！")
    else:
        messagebox.showerror("错误", "请选择功能类型！")

def run_classification():
    """运行基因匹配竖向排列功能"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    # 在新线程中运行匹配功能，显式传递关键字参数
    threading.Thread(target=run_with_progress_classification, args=(
        classify_genes_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
    )).start()

def run_gene_correspondence():
    """运行基因查询功能"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    # 在新线程中运行基因查询功能，遍历整个信息表，分别判断A列和B列是否等于当前值
    def task():
        import pandas as pd
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            df_a = pd.read_excel(file_a)
            df_b = pd.read_excel(file_b)
            if gene_column_a not in df_a.columns:
                raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
            if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
                raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")
            total = len(df_a)
            for i, (index, gene) in enumerate(df_a[gene_column_a].items()):
                if pd.isnull(gene):
                    continue
                matches = []
                for _, row in df_b.iterrows():
                    a_val = row[gene_id_column_b]
                    b_val = row[collinear_gene_column_b]
                    if a_val == gene and pd.notnull(b_val):
                        matches.append(b_val)
                    if b_val == gene and pd.notnull(a_val):
                        matches.append(a_val)
                if not matches:
                    df_a.at[index, "匹配结果"] = "无"
                else:
                    df_a.at[index, "匹配结果"] = ", ".join(map(str, matches))
                set_progress((i + 1) / total * 100)
            df_a.to_excel(output_file, index=False, engine="openpyxl")
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {output_file}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

def run_gene_search():
    """运行基因匹配横向排列功能"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    threading.Thread(target=run_with_progress, args=(
        gene_search_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
    )).start()

def run_fuzzy_match(horizontal=True):
    """运行模糊匹配功能，horizontal=True为横向模糊，False为竖向模糊"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    if horizontal:
        # 横向排列：模糊匹配和普通匹配都要各自完整实现
        # 先运行普通匹配（横向），再运行模糊匹配（横向），分别输出
        # 这里仅运行模糊匹配（横向），如需两份输出可自行扩展
        threading.Thread(target=run_with_progress_fuzzy_match, args=(
            fuzzy_match_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
        )).start()
    else:
        # 竖向排列：模糊匹配和普通匹配都要各自完整实现
        threading.Thread(target=run_with_progress_fuzzy_match, args=(
            fuzzy_match_with_progress_vertical, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
        )).start()

# 创建主窗口
root = tk.Tk()
root.title("基因工具")
root.geometry("650x550")
root.resizable(False, False)

# 设置窗口图标
root.iconbitmap(r"package\icon.ico")

# 主框架
main_frame = ttk.Notebook(root)
main_frame.pack(fill="both", expand=True)

# 功能选择页面
frame_function = ttk.LabelFrame(main_frame, text="选择功能", padding="10")
frame_function.pack(fill="x", pady=5)

# 文件选择页面
frame_files = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
frame_files.pack(fill="x", pady=5)

# 列名输入页面
frame_columns = ttk.LabelFrame(main_frame, text="列名输入", padding="10")
frame_columns.pack(fill="x", pady=5)

# 运行进度页面
frame_progress = ttk.LabelFrame(main_frame, text="运行进度", padding="10")
frame_progress.pack(fill="x", pady=5)

# 进度条和状态标签
progress_bar = ttk.Progressbar(frame_progress, orient="horizontal", length=400, mode="determinate", maximum=100)
progress_bar.pack(fill="x", padx=5, pady=5)
progress_status_label = ttk.Label(frame_progress, text="等待运行...", anchor="w")
progress_status_label.pack(fill="x", padx=5, pady=(0, 5))

def set_progress(value):
    def update():
        progress_bar["value"] = value
    root.after(0, update)

def set_progress_status(text):
    def update():
        progress_status_label.config(text=text)
    root.after(0, update)

def reset_progress():
    set_progress(0)
    set_progress_status("等待运行...")

def run_with_progress(func, *args, **kwargs):
    """带进度条和状态显示的运行函数"""
    def task():
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            func(*args, update_progress=set_progress, set_progress_status=set_progress_status)
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

def run_with_progress_gene_correspondence(func, *args):
    """带进度条和状态显示的基因查询运行函数"""
    def task():
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            func(*args, progress_callback=set_progress)
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

def run_with_progress_classification(func, *args):
    """带进度条和状态显示的基因匹配竖向排列运行函数"""
    def task():
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            func(*args, progress_callback=set_progress)
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

def highlight_cells_in_excel(file_path, highlight_cells):
    """对指定单元格进行蓝色高亮"""
    wb = load_workbook(file_path)
    ws = wb.active
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in highlight_cells:
        ws[cell].fill = blue_fill
    wb.save(file_path)

def run_with_progress_fuzzy_match(func, *args):
    """带进度条和状态显示的模糊匹配运行函数，并对不完全匹配高亮（不添加匹配类型列，支持横向/竖向多结果）"""
    def task():
        try:
            set_progress(0)
            set_progress_status("正在运行...")
            # 假设func返回(highlight_cells, output_file)
            result = func(*args, progress_callback=set_progress)
            if isinstance(result, tuple) and len(result) == 2:
                highlight_cells, output_file = result
                if highlight_cells:
                    highlight_cells_in_excel(output_file, highlight_cells)
            else:
                output_file = args[2]
            set_progress(100)
            set_progress_status("操作完成")
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {output_file}")
        except Exception as e:
            set_progress_status("运行出错")
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")
        finally:
            root.after(1000, reset_progress)
    threading.Thread(target=task).start()

# 功能选择内容
ttk.Label(frame_function, text="功能类型:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
function_combo = ttk.Combobox(
    frame_function,
    values=["基因查询", "基因匹配"],
    state="readonly",
    width=20
)
function_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# 二级下拉框：基因匹配排列方式
match_mode_label = ttk.Label(frame_function, text="排列方式:")
match_mode_combo = ttk.Combobox(
    frame_function,
    values=["横向排列", "竖向排列"],
    state="readonly",
    width=18
)
# 三级功能：模糊匹配勾选框
fuzzy_var = tk.BooleanVar(value=False)
fuzzy_check = ttk.Checkbutton(frame_function, text="启用模糊匹配", variable=fuzzy_var)
# 默认隐藏
match_mode_label.grid_remove()
match_mode_combo.grid_remove()
fuzzy_check.grid_remove()

def on_function_select(event):
    global selected_function, selected_match_mode
    selected_function = function_combo.get()
    if selected_function == "基因匹配":
        match_mode_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        match_mode_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        fuzzy_check.grid(row=0, column=4, padx=5, pady=5, sticky="w")
    else:
        match_mode_label.grid_remove()
        match_mode_combo.grid_remove()
        fuzzy_check.grid_remove()
        selected_match_mode = None

def on_match_mode_select(event):
    global selected_match_mode
    selected_match_mode = match_mode_combo.get()

function_combo.bind("<<ComboboxSelected>>", on_function_select)
match_mode_combo.bind("<<ComboboxSelected>>", on_match_mode_select)

# 文件选择内容
ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_a = ttk.Entry(frame_files, width=50)
entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_file_b = ttk.Entry(frame_files, width=50)
entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_output_file = ttk.Entry(frame_files, width=50)
entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5, sticky="w")

# 列名输入内容
ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# 运行按钮
run_button = ttk.Button(frame_progress, text="运行", command=run_function)
run_button.pack(pady=10)

root.mainloop()
root.mainloop()
