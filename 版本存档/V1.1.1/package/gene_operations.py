import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

def classify_genes_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    """基因匹配功能，支持进度更新，竖向排列，A列和B列来回查找，重复3次"""
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    # 优化映射构建
    a_to_b = {}
    b_to_a = {}
    for a_val, b_val in zip(df_b[gene_id_column_b], df_b[collinear_gene_column_b]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    result_rows = []
    total = len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        found = set()
        search_keys = {gene}
        for _ in range(3):
            # 用集合推导优化查找
            new_found = set().union(
                *(a_to_b.get(key, set()) for key in search_keys),
                *(b_to_a.get(key, set()) for key in search_keys)
            ) - found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found
        if found:
            result_rows.extend([{gene_column_a: gene, "匹配结果": match} for match in found])
        if progress_callback:
            progress_callback((i + 1) / total * 100)
    pd.DataFrame(result_rows).to_excel(output_file, index=False, engine="openpyxl")

def gene_correspondence_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    """基因对应功能，支持进度更新，遍历全表确保无遗漏"""
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    result_rows = []
    total = len(df_a)
    b_a = df_b[[gene_id_column_b, collinear_gene_column_b]].values
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        matches = [b for a, b in b_a if a == gene and pd.notnull(b)]
        matches += [a for a, b in b_a if b == gene and pd.notnull(a)]
        if matches:
            row = df_a.loc[[i]].copy()
            row["匹配结果"] = ", ".join(map(str, matches))
            result_rows.append(row)
        if progress_callback:
            progress_callback((i + 1) / total * 100)
    if result_rows:
        pd.concat(result_rows, ignore_index=True).to_excel(output_file, index=False, engine="openpyxl")
    else:
        pd.DataFrame(columns=list(df_a.columns) + ["匹配结果"]).to_excel(output_file, index=False, engine="openpyxl")

def gene_search_with_progress(fill_file, info_file, output_file, fill_col, info_a_col, info_b_col, progress_callback=None, set_progress_status=None):
    from openpyxl import load_workbook
    if set_progress_status:
        set_progress_status("正在读取填入表格...")
    wb_fill = load_workbook(fill_file)
    ws_fill = wb_fill.active

    if set_progress_status:
        set_progress_status("正在读取信息表格...")
    wb_info = load_workbook(info_file)
    ws_info = wb_info.active

    if set_progress_status:
        set_progress_status("正在查找列索引...")
    fill_col_idx = next((idx for idx, cell in enumerate(ws_fill[1], 1) if cell.value == fill_col), None)
    if fill_col_idx is None:
        raise ValueError(f"填入表格未找到列: {fill_col}")

    info_a_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_a_col), None)
    info_b_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_b_col), None)
    if info_a_idx is None or info_b_idx is None:
        raise ValueError("信息表格未找到A列或B列")

    if set_progress_status:
        set_progress_status("正在建立索引映射...")

    info_rows = list(ws_info.iter_rows(min_row=2, values_only=True))
    a_to_b = {}
    b_to_a = {}
    for row in info_rows:
        a_val = row[info_a_idx - 1]
        b_val = row[info_b_idx - 1]
        if a_val is not None and b_val is not None:
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    if set_progress_status:
        set_progress_status("正在统计最大匹配数...")

    rows = list(ws_fill.iter_rows(min_row=2))
    total = len(rows)
    all_matches = []
    max_match_count = 0
    for idx, row in enumerate(rows):
        key = row[fill_col_idx - 1].value
        found = set()
        search_keys = {key}
        for _ in range(3):
            new_found = set().union(
                *(a_to_b.get(k, set()) for k in search_keys),
                *(b_to_a.get(k, set()) for k in search_keys)
            ) - found
            if not new_found:
                break
            found.update(new_found)
            search_keys = new_found
        matches = list(found)
        if matches:
            all_matches.append(matches)
            max_match_count = max(max_match_count, len(matches))
        else:
            all_matches.append([])
        if progress_callback:
            progress_callback(int((idx + 1) / total * 50))
        if set_progress_status and idx == 0:
            set_progress_status("正在查找匹配信息...")

    if set_progress_status:
        set_progress_status("正在写入表头...")
    start_col = ws_fill.max_column + 1
    for j in range(max_match_count):
        ws_fill.cell(row=1, column=start_col + j, value=f"匹配结果{j+1}")

    if set_progress_status:
        set_progress_status("正在写入匹配结果...")
    for i, matches in enumerate(all_matches, 2):
        if not matches:
            continue
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws_fill.cell(row=i, column=start_col + j, value=val)
        if progress_callback:
            progress_callback(50 + int((i - 1) / len(rows) * 50))
    wb_fill.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)

def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None, set_progress_status=None):
    """
    横向排列：遍历全表多次查找，完全匹配和模糊匹配都支持多结果，输出多列，模糊匹配的单元格标为蓝色。
    匹配不到时不填无，直接跳过该行。
    """
    import pandas as pd
    from openpyxl import Workbook

    if set_progress_status:
        set_progress_status("正在读取表格...")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    a_to_b = {}
    b_to_a = {}
    for a_val, b_val in zip(df_b[gene_id_column_b], df_b[collinear_gene_column_b]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    all_matches = []
    all_fuzzy_flags = []
    max_match_count = 0
    total = len(df_a)
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_fuzzy_flags.append([])
            continue
        found = []
        found_flags = []
        search_keys = {gene}
        already = set()
        # 完全匹配查找
        for _ in range(3):
            new_found = set().union(
                *(a_to_b.get(k, set()) for k in search_keys),
                *(b_to_a.get(k, set()) for k in search_keys)
            ) - already
            if not new_found:
                break
            found.extend(list(new_found))
            found_flags.extend([False] * len(new_found))
            already.update(new_found)
            search_keys = new_found
        # 模糊匹配查找
        search_keys = {gene}
        already_fuzzy = set(found)
        for _ in range(3):
            fuzzy_a = {b for a in a_to_b if pd.notnull(a) and str(gene) in str(a) and a != gene for b in a_to_b[a]}
            fuzzy_b = {a for b in b_to_a if pd.notnull(b) and str(gene) in str(b) and b != gene for a in b_to_a[b]}
            new_found = (fuzzy_a | fuzzy_b) - already_fuzzy
            if not new_found:
                break
            found.extend(list(new_found))
            found_flags.extend([True] * len(new_found))
            already_fuzzy.update(new_found)
            search_keys = new_found
        if found:
            all_matches.append(found)
            all_fuzzy_flags.append(found_flags)
            max_match_count = max(max_match_count, len(found))
        else:
            all_matches.append([])
            all_fuzzy_flags.append([])
        if progress_callback:
            progress_callback((i + 1) / total * 100)

    if set_progress_status:
        set_progress_status("正在写入Excel...")

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=gene_column_a)
    for j in range(max_match_count):
        ws.cell(row=1, column=2 + j, value=f"匹配结果{j+1}")
    highlight_cells = []
    excel_row = 2
    for i, gene in enumerate(df_a[gene_column_a]):
        matches = all_matches[i]
        fuzzy_flags = all_fuzzy_flags[i]
        if not matches:
            continue
        ws.cell(row=excel_row, column=1, value=gene)
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws.cell(row=excel_row, column=2 + j, value=val)
            if j < len(fuzzy_flags) and fuzzy_flags[j]:
                col_letter = ws.cell(row=1, column=2 + j).column_letter
                highlight_cells.append(f"{col_letter}{excel_row}")
        excel_row += 1
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    return highlight_cells, output_file

def fuzzy_match_with_progress_vertical(
    file_a, file_b, output_file,
    gene_column_a, gene_id_column_b, collinear_gene_column_b,
    progress_callback=None, set_progress_status=None
):
    """
    竖向模糊匹配，A列和B列来回查找，重复3次，完全匹配和原有逻辑不变，仅对不完全匹配部分用蓝色高亮，不添加匹配类型列，支持竖向多结果。
    """
    import pandas as pd
    from openpyxl import Workbook

    if set_progress_status:
        set_progress_status("正在读取表格...")
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    if gene_column_a not in df_a.columns:
        raise ValueError(f"表格 A 中不存在列名 '{gene_column_a}'")
    if gene_id_column_b not in df_b.columns or collinear_gene_column_b not in df_b.columns:
        raise ValueError(f"表格 B 中不存在列名 '{gene_id_column_b}' 或 '{collinear_gene_column_b}'")

    if set_progress_status:
        set_progress_status("正在查找匹配结果...")

    a_to_b = {}
    b_to_a = {}
    for a_val, b_val in zip(df_b[gene_id_column_b], df_b[collinear_gene_column_b]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    result_rows = []
    highlight_cells = []
    total = len(df_a)
    excel_row = 2

    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        found = set()
        found_type = dict()
        search_keys = {gene}
        for round_idx in range(3):
            if set_progress_status:
                set_progress_status(f"第{round_idx+1}/3轮查找：正在处理第 {i+1}/{total} 个基因...")
            # 完全匹配
            new_found = set().union(
                *(a_to_b.get(k, set()) for k in search_keys),
                *(b_to_a.get(k, set()) for k in search_keys)
            ) - found
            for m in new_found:
                found_type[m] = False
            # 模糊匹配
            fuzzy_a = {b for a in a_to_b if pd.notnull(a) and str(k) in str(a) and a != k for k in search_keys for b in a_to_b[a]}
            fuzzy_b = {a for b in b_to_a if pd.notnull(b) and str(k) in str(b) and b != k for k in search_keys for a in b_to_a[b]}
            fuzzy_new = (fuzzy_a | fuzzy_b) - found
            for m in fuzzy_new:
                found_type[m] = True
            all_new = new_found | fuzzy_new
            if not all_new:
                break
            found.update(all_new)
            search_keys = all_new
        matches = list(found)
        if matches:
            for m in matches:
                result_rows.append([gene, m])
                if found_type.get(m, False):
                    highlight_cells.append(f"B{excel_row}")
                excel_row += 1
        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"正在处理第 {i+1}/{total} 个基因...")

    if set_progress_status:
        set_progress_status("正在写入Excel...")

    wb = Workbook()
    ws = wb.active
    ws.append([gene_column_a, "匹配结果"])
    for row in result_rows:
        ws.append(row)
    wb.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    return highlight_cells, output_file



