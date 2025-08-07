import tkinter as tk
from tkinter import ttk, filedialog, messagebox, LabelFrame
from package.gene_operations import classify_genes_with_progress, gene_correspondence_with_progress
import threading
from openpyxl import load_workbook

# 全局变量
selected_function = None

def select_file(entry):
    """打开文件选择对话框并设置文件路径"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def select_output_file(entry):
    """选择输出文件路径"""
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

def update_column_options(entry, file_path):
    """更新列名下拉选项"""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        columns = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        entry["values"] = columns
    except Exception as e:
        messagebox.showerror("错误", f"无法读取列名：{str(e)}")

def select_file_and_update_columns(entry_file, *entry_columns):
    """选择文件并更新多个列名选项"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)
        for entry_column in entry_columns:
            update_column_options(entry_column, file_path)

def run_function():
    """运行所选功能"""
    if selected_function == "基因匹配":
        run_gene_correspondence()
    elif selected_function == "基因对应":
        run_classification()
    else:
        messagebox.showerror("错误", "请选择功能类型！")

def run_classification():
    """运行基因匹配功能"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    # 在新线程中运行匹配功能
    threading.Thread(target=run_with_progress, args=(classify_genes_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b)).start()

def run_gene_correspondence():
    """运行基因对应功能"""
    file_a = entry_file_a.get()
    file_b = entry_file_b.get()
    output_file = entry_output_file.get()
    gene_column_a = entry_gene_column_a.get()
    gene_id_column_b = entry_gene_id_column_b.get()
    collinear_gene_column_b = entry_collinear_gene_column_b.get()

    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")
        return

    # 在新线程中运行基因对应功能
    threading.Thread(target=run_with_progress, args=(gene_correspondence_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b)).start()

def run_with_progress(func, *args):
    """带进度条的运行逻辑"""
    try:
        progress_bar["value"] = 0
        progress_label.config(text="运行中...")
        func(*args, update_progress)
        progress_label.config(text="完成")
        messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")
    except Exception as e:
        progress_label.config(text="出错")
        messagebox.showerror("错误", f"运行时出现错误：{str(e)}")

def update_progress(value):
    """更新进度条"""
    progress_bar["value"] = value
    root.update_idletasks()

# 创建主窗口
root = tk.Tk()
root.title("基因工具")
root.geometry("650x550")
root.resizable(False, False)

# 设置窗口图标
root.iconbitmap("package\icon.ico")

# 主框架
main_frame = ttk.Notebook(root)
main_frame.pack(fill="both", expand=True)

# 功能选择页面
frame_function = ttk.LabelFrame(main_frame, text="选择功能", padding="10")
frame_function.pack(fill="x", pady=5)

# 文件选择页面
frame_files = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
frame_files.pack(fill="x", pady=5)

# 列名输入页面
frame_columns = ttk.LabelFrame(main_frame, text="列名输入", padding="10")
frame_columns.pack(fill="x", pady=5)

# 运行进度页面
frame_progress = ttk.LabelFrame(main_frame, text="运行进度", padding="10")
frame_progress.pack(fill="x", pady=5)

# 功能选择内容
ttk.Label(frame_function, text="功能类型:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
function_combo = ttk.Combobox(frame_function, values=["基因匹配", "基因对应"], state="readonly", width=20)
function_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

def on_function_select(event):
    global selected_function
    selected_function = function_combo.get()

function_combo.bind("<<ComboboxSelected>>", on_function_select)

# 文件选择内容
ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_a = ttk.Entry(frame_files, width=50)
entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_file_b = ttk.Entry(frame_files, width=50)
entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_output_file = ttk.Entry(frame_files, width=50)
entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5, sticky="w")

# 列名输入内容
ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# 运行进度内容
progress_bar = ttk.Progressbar(frame_progress, orient="horizontal", length=400, mode="determinate")
progress_bar.pack(fill="x", padx=5, pady=5)
progress_label = ttk.Label(frame_progress, text="")
progress_label.pack()

# 运行按钮
run_button = ttk.Button(frame_progress, text="运行", command=run_function)
run_button.pack(pady=10)

root.mainloop()
