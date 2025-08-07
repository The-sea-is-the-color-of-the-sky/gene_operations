import tkinter as tk  # 导入tkinter库，用于GUI开发
from tkinter import ttk, filedialog, messagebox  # 导入ttk、文件对话框和消息框
from package.gene_operations import (  # 从自定义包导入基因操作相关函数
    classify_genes_with_progress,
    gene_search_with_progress,
    fuzzy_match_with_progress,
    gene_correspondence_with_progress
)
import threading  # 导入线程库，实现多线程
from openpyxl import load_workbook  # 导入openpyxl用于操作Excel文件
from openpyxl.styles import PatternFill  # 导入单元格填充样式

selected_function = None  # 当前选择的功能类型
selected_match_mode = None  # 当前选择的匹配排列方式

def select_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])  # 弹出文件选择框，选择Excel文件
    if file_path:
        entry.delete(0, tk.END)  # 清空输入框
        entry.insert(0, file_path)  # 插入选择的文件路径

def select_output_file(entry):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])  # 弹出保存文件对话框
    if file_path:
        entry.delete(0, tk.END)  # 清空输入框
        entry.insert(0, file_path)  # 插入保存的文件路径

def update_column_options(entry, file_path):
    try:
        workbook = load_workbook(file_path, read_only=True)  # 只读方式打开Excel文件
        sheet = workbook.active  # 获取活动表
        columns = [cell.value for cell in next(sheet.iter_rows(max_row=1))]  # 读取第一行作为列名
        entry["values"] = columns  # 设置下拉框的可选值
    except Exception as e:
        messagebox.showerror("错误", f"无法读取列名：{str(e)}")  # 弹出错误提示

def select_file_and_update_columns(entry_file, *entry_columns):
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])  # 选择Excel文件
    if file_path:
        entry_file.delete(0, tk.END)  # 清空输入框
        entry_file.insert(0, file_path)  # 插入文件路径
        for entry_column in entry_columns:
            update_column_options(entry_column, file_path)  # 更新列名下拉框

def run_function():
    if selected_function == "基因查询":  # 判断选择的功能
        run_gene_correspondence()  # 运行基因查询
    elif selected_function == "基因匹配":
        is_fuzzy = fuzzy_var.get()  # 获取是否启用模糊匹配
        if selected_match_mode == "横向排列":
            if is_fuzzy:
                run_fuzzy_match(horizontal=True)  # 横向模糊匹配
            else:
                run_gene_search()  # 横向精确匹配
        elif selected_match_mode == "竖向排列":
            if is_fuzzy:
                run_fuzzy_match(horizontal=False)  # 竖向模糊匹配
            else:
                run_classification()  # 竖向精确匹配
        else:
            messagebox.showerror("错误", "请选择基因匹配的排列方式！")  # 未选择排列方式提示
    else:
        messagebox.showerror("错误", "请选择功能类型！")  # 未选择功能类型提示

def run_classification():
    file_a = entry_file_a.get()  # 获取表格A路径
    file_b = entry_file_b.get()  # 获取表格B路径
    output_file = entry_output_file.get()  # 获取输出文件路径
    gene_column_a = entry_gene_column_a.get()  # 获取A表目标列名
    gene_id_column_b = entry_gene_id_column_b.get()  # 获取B表A列名
    collinear_gene_column_b = entry_collinear_gene_column_b.get()  # 获取B表B列名
    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")  # 检查必填项
        return
    threading.Thread(target=run_with_progress_classification, args=(
        classify_genes_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
    )).start()  # 启动新线程运行分类

def run_gene_correspondence():
    file_a = entry_file_a.get()  # 获取表格A路径
    file_b = entry_file_b.get()  # 获取表格B路径
    output_file = entry_output_file.get()  # 获取输出文件路径
    gene_column_a = entry_gene_column_a.get()  # 获取A表目标列名
    gene_id_column_b = entry_gene_id_column_b.get()  # 获取B表A列名
    collinear_gene_column_b = entry_collinear_gene_column_b.get()  # 获取B表B列名
    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")  # 检查必填项
        return
    threading.Thread(target=run_with_progress, args=(
        gene_correspondence_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
    )).start()  # 启动新线程运行

def run_gene_search():
    file_a = entry_file_a.get()  # 获取表格A路径
    file_b = entry_file_b.get()  # 获取表格B路径
    output_file = entry_output_file.get()  # 获取输出文件路径
    gene_column_a = entry_gene_column_a.get()  # 获取A表目标列名
    gene_id_column_b = entry_gene_id_column_b.get()  # 获取B表A列名
    collinear_gene_column_b = entry_collinear_gene_column_b.get()  # 获取B表B列名
    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")  # 检查必填项
        return
    threading.Thread(target=run_with_progress, args=(
        gene_search_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
    )).start()  # 启动新线程运行

def run_fuzzy_match(horizontal=True):
    file_a = entry_file_a.get()  # 获取表格A路径
    file_b = entry_file_b.get()  # 获取表格B路径
    output_file = entry_output_file.get()  # 获取输出文件路径
    gene_column_a = entry_gene_column_a.get()  # 获取A表目标列名
    gene_id_column_b = entry_gene_id_column_b.get()  # 获取B表A列名
    collinear_gene_column_b = entry_collinear_gene_column_b.get()  # 获取B表B列名
    if not all([file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b]):
        messagebox.showerror("错误", "请填写所有字段！")  # 检查必填项
        return
    if horizontal:
        threading.Thread(target=run_with_progress, args=(
            fuzzy_match_with_progress, file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
        )).start()  # 横向模糊匹配新线程
    else:
        threading.Thread(target=run_with_progress_fuzzy_match, args=(
            file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b
        )).start()  # 竖向模糊匹配新线程

root = tk.Tk()  # 创建主窗口
root.title("基因工具")  # 设置窗口标题
root.geometry("650x550")  # 设置窗口大小
root.resizable(False, False)  # 禁止调整窗口大小
root.iconbitmap(r"package\icon.ico")  # 设置窗口图标

main_frame = ttk.Notebook(root)  # 创建多标签页控件
main_frame.pack(fill="both", expand=True)  # 填充整个窗口

frame_function = ttk.LabelFrame(main_frame, text="选择功能", padding="10")  # 功能选择区域
frame_function.pack(fill="x", pady=5)
frame_files = ttk.LabelFrame(main_frame, text="文件选择", padding="10")  # 文件选择区域
frame_files.pack(fill="x", pady=5)
frame_columns = ttk.LabelFrame(main_frame, text="列名输入", padding="10")  # 列名输入区域
frame_columns.pack(fill="x", pady=5)
frame_progress = ttk.LabelFrame(main_frame, text="运行进度", padding="10")  # 进度显示区域
frame_progress.pack(fill="x", pady=5)

progress_bar = ttk.Progressbar(frame_progress, orient="horizontal", length=400, mode="determinate", maximum=100)  # 进度条
progress_bar.pack(fill="x", padx=5, pady=5)
progress_status_label = ttk.Label(frame_progress, text="等待运行...", anchor="w")  # 进度状态标签
progress_status_label.pack(fill="x", padx=5, pady=(0, 5))

def set_progress(value):
    def update():
        progress_bar["value"] = value  # 设置进度条的值
    root.after(0, update)  # 在主线程中更新

def set_progress_status(text):
    def update():
        progress_status_label.config(text=text)  # 设置进度状态文本
    root.after(0, update)  # 在主线程中更新

def reset_progress():
    set_progress(0)  # 重置进度条
    set_progress_status("等待运行...")  # 重置状态文本

def run_with_progress(func, *args, **kwargs):
    def task():
        try:
            set_progress(0)  # 初始化进度
            set_progress_status("正在运行...")  # 设置状态
            func(*args, progress_callback=set_progress, set_progress_status=set_progress_status)  # 调用目标函数
            set_progress(100)  # 设置进度为100%
            set_progress_status("操作完成")  # 设置状态为完成
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")  # 弹出成功提示
        except Exception as e:
            set_progress_status("运行出错")  # 设置状态为出错
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")  # 弹出错误提示
        finally:
            root.after(1000, reset_progress)  # 1秒后重置进度
    threading.Thread(target=task).start()  # 启动新线程

def run_with_progress_classification(func, *args):
    def task():
        try:
            set_progress(0)  # 初始化进度
            set_progress_status("正在运行...")  # 设置状态
            func(*args, progress_callback=set_progress)  # 调用目标函数
            set_progress(100)  # 设置进度为100%
            set_progress_status("操作完成")  # 设置状态为完成
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {args[2]}")  # 弹出成功提示
        except Exception as e:
            set_progress_status("运行出错")  # 设置状态为出错
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")  # 弹出错误提示
        finally:
            root.after(1000, reset_progress)  # 1秒后重置进度
    threading.Thread(target=task).start()  # 启动新线程

def highlight_cells_in_excel(file_path, highlight_cells):
    wb = load_workbook(file_path)  # 打开Excel文件
    ws = wb.active  # 获取活动表
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 设置蓝色填充
    for cell in highlight_cells:
        ws[cell].fill = blue_fill  # 高亮指定单元格
    wb.save(file_path)  # 保存文件

def run_with_progress_fuzzy_match(func, *args):
    def task():
        try:
            set_progress(0)  # 初始化进度
            set_progress_status("正在运行...")  # 设置状态
            result = func(*args, progress_callback=set_progress)  # 调用目标函数
            if isinstance(result, tuple) and len(result) == 2:
                highlight_cells, output_file = result  # 获取高亮单元格和输出文件
                if highlight_cells:
                    highlight_cells_in_excel(output_file, highlight_cells)  # 高亮单元格
            else:
                output_file = args[2]
            set_progress(100)  # 设置进度为100%
            set_progress_status("操作完成")  # 设置状态为完成
            messagebox.showinfo("成功", f"操作完成，结果已保存到 {output_file}")  # 弹出成功提示
        except Exception as e:
            set_progress_status("运行出错")  # 设置状态为出错
            messagebox.showerror("错误", f"运行时出现错误：{str(e)}")  # 弹出错误提示
        finally:
            root.after(1000, reset_progress)  # 1秒后重置进度
    threading.Thread(target=task).start()  # 启动新线程

ttk.Label(frame_function, text="功能类型:").grid(row=0, column=0, padx=5, pady=5, sticky="w")  # 功能类型标签
function_combo = ttk.Combobox(
    frame_function,
    values=["基因查询", "基因匹配"],
    state="readonly",
    width=20
)  # 功能类型下拉框
function_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

match_mode_label = ttk.Label(frame_function, text="排列方式:")  # 排列方式标签
match_mode_combo = ttk.Combobox(
    frame_function,
    values=["横向排列", "竖向排列"],
    state="readonly",
    width=18
)  # 排列方式下拉框
# 三级功能：模糊匹配勾选框
fuzzy_var = tk.BooleanVar(value=False)  # 模糊匹配变量
fuzzy_check = ttk.Checkbutton(frame_function, text="启用模糊匹配", variable=fuzzy_var)  # 模糊匹配勾选框
# 默认隐藏
match_mode_label.grid_remove()  # 隐藏排列方式标签
match_mode_combo.grid_remove()  # 隐藏排列方式下拉框
fuzzy_check.grid_remove()  # 隐藏模糊匹配勾选框

def on_function_select(event):
    global selected_function, selected_match_mode
    selected_function = function_combo.get()  # 获取当前选择的功能
    if selected_function == "基因匹配":
        match_mode_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")  # 显示排列方式标签
        match_mode_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")  # 显示排列方式下拉框
        fuzzy_check.grid(row=0, column=4, padx=5, pady=5, sticky="w")  # 显示模糊匹配勾选框
    else:
        match_mode_label.grid_remove()  # 隐藏排列方式标签
        match_mode_combo.grid_remove()  # 隐藏排列方式下拉框
        fuzzy_check.grid_remove()  # 隐藏模糊匹配勾选框
        selected_match_mode = None  # 重置排列方式

def on_match_mode_select(event):
    global selected_match_mode
    selected_match_mode = match_mode_combo.get()  # 获取当前选择的排列方式

function_combo.bind("<<ComboboxSelected>>", on_function_select)  # 绑定功能选择事件
match_mode_combo.bind("<<ComboboxSelected>>", on_match_mode_select)  # 绑定排列方式选择事件

# 文件选择内容
ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")  # 表格A标签
entry_file_a = ttk.Entry(frame_files, width=50)  # 表格A路径输入框
entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5, sticky="w")  # 选择文件按钮

ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")  # 表格B标签
entry_file_b = ttk.Entry(frame_files, width=50)  # 表格B路径输入框
entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5, sticky="w")  # 选择文件按钮

ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")  # 输出文件标签
entry_output_file = ttk.Entry(frame_files, width=50)  # 输出文件路径输入框
entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5, sticky="w")  # 选择路径按钮

# 列名输入内容
ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")  # A表目标列名标签
entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")  # A表目标列名下拉框
entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")  # B表A列名标签
entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")  # B表A列名下拉框
entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")  # B表B列名标签
entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")  # B表B列名下拉框
entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# 运行按钮
run_button = ttk.Button(frame_progress, text="运行", command=run_function)  # 运行按钮
run_button.pack(pady=10)

fuzzy_check.grid_remove()  # 默认隐藏模糊匹配勾选框
match_mode_label.grid_remove()
match_mode_combo.grid_remove()  # 默认隐藏排列方式标签和下拉框


def on_function_select(event):
    global selected_function, selected_match_mode
    selected_function = function_combo.get()
    if selected_function == "基因匹配":
        match_mode_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        match_mode_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        fuzzy_check.grid(row=0, column=4, padx=5, pady=5, sticky="w")
    else:
        match_mode_label.grid_remove()
        match_mode_combo.grid_remove()
        fuzzy_check.grid_remove()
        selected_match_mode = None

def on_match_mode_select(event):
    global selected_match_mode
    selected_match_mode = match_mode_combo.get()

function_combo.bind("<<ComboboxSelected>>", on_function_select)
match_mode_combo.bind("<<ComboboxSelected>>", on_match_mode_select)

# 文件选择内容
ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_a = ttk.Entry(frame_files, width=50)
entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_file_b = ttk.Entry(frame_files, width=50)
entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_output_file = ttk.Entry(frame_files, width=50)
entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5, sticky="w")

# 列名输入内容
ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# 运行按钮
run_button = ttk.Button(frame_progress, text="运行", command=run_function)
run_button.pack(pady=10)

root.mainloop()

# 文件选择内容
ttk.Label(frame_files, text="填入的表格文件路径:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_file_a = ttk.Entry(frame_files, width=50)
entry_file_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_a, entry_gene_column_a)).grid(row=0, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="基因信息表格文件路径:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_file_b = ttk.Entry(frame_files, width=50)
entry_file_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择文件", command=lambda: select_file_and_update_columns(entry_file_b, entry_gene_id_column_b, entry_collinear_gene_column_b)).grid(row=1, column=2, padx=5, pady=5, sticky="w")

ttk.Label(frame_files, text="输出文件路径:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_output_file = ttk.Entry(frame_files, width=50)
entry_output_file.grid(row=2, column=1, padx=5, pady=5, sticky="w")
ttk.Button(frame_files, text="选择路径", command=lambda: select_output_file(entry_output_file)).grid(row=2, column=2, padx=5, pady=5, sticky="w")

# 列名输入内容
ttk.Label(frame_columns, text="填入表格目标列名:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_gene_column_a = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_column_a.grid(row=0, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的A列名:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_gene_id_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_gene_id_column_b.grid(row=1, column=1, padx=5, pady=5, sticky="w")

ttk.Label(frame_columns, text="信息表格的B列名:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_collinear_gene_column_b = ttk.Combobox(frame_columns, width=40, state="readonly")
entry_collinear_gene_column_b.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# 运行按钮
run_button = ttk.Button(frame_progress, text="运行", command=run_function)
run_button.pack(pady=10)

root.mainloop()
root.mainloop()
root.mainloop()
