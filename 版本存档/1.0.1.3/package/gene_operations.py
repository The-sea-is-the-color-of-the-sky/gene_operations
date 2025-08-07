import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# ======= 通用工具函数 =======
def validate_columns(df, *columns):
    for col in columns:
        if col not in df.columns:
            raise ValueError(f"表格中不存在列名 '{col}'")

def build_bidirectional_map(df, col1, col2):
    a_to_b, b_to_a = {}, {}
    for a_val, b_val in zip(df[col1], df[col2]):
        if pd.notnull(a_val) and pd.notnull(b_val):
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)
    return a_to_b, b_to_a

def recursive_search(keys, a_to_b, b_to_a, depth=3, fuzzy=False):
    found = set(keys)
    current = set(keys)
    for _ in range(depth):
        next_found = set()
        for key in current:
            if fuzzy:
                next_found.update(b for a in a_to_b if key in str(a) and a != key for b in a_to_b[a])
                next_found.update(a for b in b_to_a if key in str(b) and b != key for a in b_to_a[b])
            else:
                next_found |= a_to_b.get(key, set())
                next_found |= b_to_a.get(key, set())
        new = next_found - found
        if not new:
            break
        found |= new
        current = new
    return found - set(keys)

# ======= 功能函数 =======

def classify_genes_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None):
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)
    validate_columns(df_a, gene_column_a)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b)
    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b)

    results = []
    total = len(df_a)

    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            continue
        found = recursive_search({gene}, a_to_b, b_to_a, depth=3)
        for match in found:
            results.append({gene_column_a: gene, "匹配结果": match})
        if progress_callback:
            progress_callback((i + 1) / total * 100)

    pd.DataFrame(results).to_excel(output_file, index=False, engine="openpyxl")

def gene_search_with_progress(fill_file, info_file, output_file, fill_col, info_a_col, info_b_col, progress_callback=None, set_progress_status=None):
    if set_progress_status:
        set_progress_status("读取填入表...")

    wb_fill = load_workbook(fill_file)
    ws_fill = wb_fill.active

    if set_progress_status:
        set_progress_status("读取信息表...")

    wb_info = load_workbook(info_file)
    ws_info = wb_info.active

    fill_idx = next((idx for idx, cell in enumerate(ws_fill[1], 1) if cell.value == fill_col), None)
    info_a_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_a_col), None)
    info_b_idx = next((idx for idx, cell in enumerate(ws_info[1], 1) if cell.value == info_b_col), None)

    if None in (fill_idx, info_a_idx, info_b_idx):
        raise ValueError("缺少必要列")

    info_rows = list(ws_info.iter_rows(min_row=2, values_only=True))

    a_to_b, b_to_a = {}, {}
    for row in info_rows:
        a_val = row[info_a_idx - 1]
        b_val = row[info_b_idx - 1]
        if a_val is not None and b_val is not None:
            a_to_b.setdefault(a_val, set()).add(b_val)
            b_to_a.setdefault(b_val, set()).add(a_val)

    rows = list(ws_fill.iter_rows(min_row=2))
    max_match_count = 0
    all_matches = []

    total = len(rows)
    for idx, row in enumerate(rows):
        key = row[fill_idx - 1].value
        found = recursive_search({key}, a_to_b, b_to_a, depth=3) if key else set()
        matches = list(found)
        all_matches.append(matches)
        max_match_count = max(max_match_count, len(matches))
        if progress_callback:
            progress_callback(int((idx + 1) / total * 50))

    start_col = ws_fill.max_column + 1
    for j in range(max_match_count):
        ws_fill.cell(row=1, column=start_col + j, value=f"匹配结果{j+1}")

    for i, matches in enumerate(all_matches, 2):
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            ws_fill.cell(row=i, column=start_col + j, value=val)
        if progress_callback:
            progress_callback(50 + int((i - 1) / total * 50))

    wb_fill.save(output_file)
    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)

def fuzzy_match_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None, set_progress_status=None):
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    validate_columns(df_a, gene_column_a)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b)

    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b)

    all_matches = []
    all_flags = []
    max_match_count = 0
    total = len(df_a)

    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            all_matches.append([])
            all_flags.append([])
            continue

        exact = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=False)
        fuzzy = recursive_search({gene}, a_to_b, b_to_a, depth=3, fuzzy=True)

        exact_list = list(exact)
        fuzzy_list = list(fuzzy - exact)
        matches = exact_list + fuzzy_list
        flags = [False] * len(exact_list) + [True] * len(fuzzy_list)

        all_matches.append(matches)
        all_flags.append(flags)
        max_match_count = max(max_match_count, len(matches))

        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"处理进度：{i + 1}/{total}")

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value=gene_column_a)
    for j in range(max_match_count):
        ws.cell(row=1, column=2 + j, value=f"匹配结果{j+1}")

    blue_fill = PatternFill(fill_type="solid", fgColor="87CEEB")

    for i, gene in enumerate(df_a[gene_column_a], 2):
        ws.cell(row=i, column=1, value=gene)
        matches = all_matches[i - 2]
        flags = all_flags[i - 2]
        for j in range(max_match_count):
            val = matches[j] if j < len(matches) else ""
            cell = ws.cell(row=i, column=2 + j, value=val)
            if j < len(flags) and flags[j]:
                cell.fill = blue_fill

    wb.save(output_file)

    if set_progress_status:
        set_progress_status("保存完成")

    highlight_cells = []
    for i, flags in enumerate(all_flags, 2):
        for j, flag in enumerate(flags):
            if flag:
                col_letter = ws.cell(row=1, column=2 + j).column_letter
                highlight_cells.append(f"{col_letter}{i}")

    return highlight_cells, output_file

def gene_correspondence_with_progress(file_a, file_b, output_file, gene_column_a, gene_id_column_b, collinear_gene_column_b, progress_callback=None, set_progress_status=None):
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    validate_columns(df_a, gene_column_a)
    validate_columns(df_b, gene_id_column_b, collinear_gene_column_b)

    a_to_b, b_to_a = build_bidirectional_map(df_b, gene_id_column_b, collinear_gene_column_b)
    total = len(df_a)

    match_results = []
    for i, gene in enumerate(df_a[gene_column_a]):
        if pd.isnull(gene):
            match_results.append("无")
            if progress_callback:
                progress_callback((i + 1) / total * 100)
            continue

        matches = set()
        matches |= a_to_b.get(gene, set())
        matches |= b_to_a.get(gene, set())
        matches.discard(gene)

        if matches:
            match_results.append(", ".join(map(str, matches)))
        else:
            match_results.append("无")

        if progress_callback:
            progress_callback((i + 1) / total * 100)
        if set_progress_status and (i % 10 == 0 or i == total - 1):
            set_progress_status(f"处理进度：{i + 1}/{total}")

    df_a["匹配结果"] = match_results
    df_a.to_excel(output_file, index=False, engine="openpyxl")

    if set_progress_status:
        set_progress_status("保存完成")
    if progress_callback:
        progress_callback(100)
